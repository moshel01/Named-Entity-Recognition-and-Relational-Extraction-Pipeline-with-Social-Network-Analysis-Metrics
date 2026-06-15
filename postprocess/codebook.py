# Auto-generated SNA codebook: documents every node/edge variable, type,
# tie class, and evidence tier of a run so a reader who has never seen the
# pipeline can interpret the Gephi tables. Modeled on conventional network
# codebooks (boundary specification + variable definitions + coding rules),
# filled with this run's actual value inventories and counts.

from __future__ import annotations

import logging
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from . import tie_classes
from .evidence_tiers import TIER_DOCS

logger = logging.getLogger(__name__)

# Variable definitions for the standard node columns.
_NODE_DEFS: dict[str, str] = {
    "Id": "Stable entity id (hash of normalized name + type). Joins to edge Source/Target.",
    "Label": "Canonical display name after deduplication / alias folding.",
    "type": "Entity type (see the Entity Types sheet).",
    "mention_count": "Total surface mentions across the corpus (all documents, all chunks).",
    "doc_count": "Number of distinct documents the entity appears in.",
    "aliases": "Alternative surface forms folded into this node (semicolon-separated).",
    "first_year": "Earliest year attached to any edge or mention of this node.",
    "last_year": "Latest year attached to any edge or mention of this node.",
    "confidence": "Best extraction confidence across mentions, 0-1.",
    "sna_constraint": "Burt's constraint: how redundant the node's contacts are (low = brokerage position).",
    "sna_effective_size": "Effective size of the ego network (non-redundant contacts).",
    "sna_is_articulation": "True if removing the node disconnects part of the graph.",
    "tag_subtype": "Fine-grained subtype assigned by enrichment (see Entity Types sheet).",
    "tag_entity_scope": "specific = a concrete named entity; generic = a category reference.",
    "tag_relevance_tier": "core / secondary / peripheral salience tier from degree + mentions.",
    "tag_degree": "Total degree across all tie classes.",
    "tag_reference_figure": "True for public/historical figures mentioned but outside the studied population.",
    "attr_is_author": "True if this node is the detected author/narrator of a document.",
    "attr_author_doc": "Document id where this node was detected as the author.",
    "attr_narrator": "True if the node was built from first-person narration.",
    "attr_surface_pronoun": "The pronoun the narrator used (e.g. 'ich').",
    "attr_propn_ratio": "Share of mentions POS-tagged as proper noun, 0-1. Low values suggest a category word.",
    "attr_suspect_common_noun": "True when propn_ratio < 0.5 for a proper-name type: possible NER noise, kept for filtering.",
    "attr_evidence": "Sample sentence supporting the entity.",
    "attr_evidence_doc": "Document the evidence sentence came from.",
    "attr_source": "Set to 'metadata' when the node was created from the metadata spreadsheet, not the text.",
}
_NODE_PREFIX_DEFS: list[tuple[str, str]] = [
    ("deg_", "Degree within one tie class (edges of that class touching this node)."),
    ("attr_", "Domain attribute merged from enrichment or the metadata spreadsheet."),
    ("sna_", "Precomputed NetworkX metric (see graph_report.json for graph-level QA)."),
    ("tag_", "Pipeline-assigned tag for filtering in Gephi."),
]

_EDGE_DEFS: dict[str, str] = {
    "Id": "Stable edge id (hash of endpoints + relation type).",
    "Source": "Entity id of the source node.",
    "Target": "Entity id of the target node.",
    "Type": "Directed when the relation is asymmetric (works_for); Undirected for mutual ties (met_with).",
    "Label": "Relation type (same as rel_type; Gephi display).",
    "rel_type": "Specific relation verb (see the Relation Inventory sheet).",
    "tie_class": "Social-tie class of the relation (see the Tie Classes sheet).",
    "connection_type": "Physical / ideological / organizational / biographical axis, orthogonal to tie_class: separates a direct material tie (meeting, funding, combat, kinship) from a shared/opposed-belief one.",
    "polarity": "Sign of the tie: positive / negative / neutral (for signed-network analysis).",
    "Weight": "Number of distinct documents supporting the edge (corroboration), not raw mentions.",
    "n_mentions": "Raw number of supporting relationship mentions.",
    "n_sources": "Number of distinct letters/documents (by letter id) supporting the edge.",
    "reciprocal": "True if the relation was asserted in both directions.",
    "suspect_membership": "True for membership edges whose target is not an org/institution: possible extraction error, kept for filtering.",
    "evidence_unverified": "True when the LLM's evidence string is not a verbatim span of the source text: possible paraphrase or hallucination, kept for filtering.",
    "period": "Domain-defined historical period of the edge's year.",
    "year": "Year extracted from the evidence, if any.",
    "origin": "extracted (stated in text) / inferred (added by the pipeline).",
    "edge_source": "Provenance of the edge (see the Evidence Tiers sheet).",
    "confidence": "Best extraction confidence across supporting mentions, 0-1.",
    "source_name": "Display name of the source node.",
    "target_name": "Display name of the target node.",
    "letter_id": "Letter/document identifier of a supporting document.",
    "evidence": "Sample verbatim text span supporting the edge (truncated to 500 chars).",
    "is_bridge": "True if removing the edge disconnects part of the graph (NetworkX).",
}

_TIE_CLASS_DEFS: dict[str, str] = {
    "interaction": "Person-to-person social tie actually narrated. The headline social network.",
    "affiliation": "Person-to-organization/institution tie (membership, employment, schooling). Two-mode.",
    "participation": "Person-to-event tie (fought in, attended). Two-mode.",
    "biographical": "Person-to-place or person-to-rank tie (born in, resided in, promoted to). Attribute-like.",
    "stance": "Attitude or opinion (supported, opposed). Discourse signal, not a social tie.",
    "cooccurrence": "Co-presence in the same context only. Weakest evidence layer.",
    "other": "Unclassifiable relations and dedup artifacts.",
}

def _col_def(col: str, defs: dict[str, str], prefixes: list[tuple[str, str]] | None = None) -> str:
    if col in defs:
        return defs[col]
    for pre, text in (prefixes or []):
        if col.startswith(pre):
            return text
    return "Domain- or run-specific attribute."


def write_codebook(
    run_dir: Path,
    tables: Any,
    config: Any,
    domain: Any = None,
    model: str = "",
) -> Path | None:
    """Write codebook.xlsx into the run dir. Fail-soft: returns None on error."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
    except ImportError:
        logger.warning("openpyxl not available; skipping codebook export.")
        return None
    try:
        return _write(run_dir, tables, config, domain, model, Workbook, Font, Alignment)
    except Exception as exc:  # noqa: BLE001 - never fail a run over documentation
        logger.warning("Codebook export failed: %s", exc)
        return None


def _write(run_dir, tables, config, domain, model, Workbook, Font, Alignment):
    nodes: list[dict] = tables.nodes
    edges: list[dict] = tables.edges
    wb = Workbook()
    bold = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    def sheet(title, header, rows, widths):
        ws = wb.create_sheet(title)
        ws.append(header)
        for c in ws[1]:
            c.font = bold
        for row in rows:
            ws.append(row)
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + i) if i <= 26 else "A"].width = w
            for cell in list(ws.columns)[i - 1]:
                cell.alignment = wrap
        return ws

    # 1. Overview: provenance + boundary specification.
    type_counts = Counter(n.get("type", "") for n in nodes)
    class_counts = Counter(e.get("tie_class", "") for e in edges)
    overview = [
        ("What this file is",
         "Codebook for the extracted social network: defines every variable in "
         "gephi_nodes.csv / gephi_edges.csv and the value inventories of this run."),
        ("Run name", getattr(config, "run_name", "")),
        ("Generated", datetime.now(timezone.utc).isoformat(timespec="seconds")),
        ("Extraction mode", getattr(config, "mode", "")),
        ("Model", model),
        ("Domain", getattr(getattr(config, "domain", None), "name", "")),
        ("Nodes", len(nodes)),
        ("Edges", len(edges)),
        ("Node boundary",
         "Every named entity of the configured types surviving quality review "
         "(stopword, junk-name, and proper-noun gates). Suspected noise is kept "
         "and tagged (attr_suspect_common_noun), not silently removed."),
        ("Edge boundary",
         "Every relation stated in the text plus inferred layers, each labeled by "
         "edge_source and tie_class. A mention is not a tie: filter on tie_class "
         "and the evidence tiers (see those sheets) to choose your network."),
        ("Multi-view files",
         "network.gexf = full graph; graph_interaction/affiliation/discourse.gexf "
         "= single-layer views."
         # Dynamic graph only exists when the corpus yielded dated events.
         + (" network_dynamic.gexf adds the time axis."
            if (Path(run_dir) / "network_dynamic.gexf").exists() else "")),
    ]
    note = getattr(getattr(config, "export", None), "codebook_note", "")
    if note:
        overview.append(("Corpus note", note))
    sheet("Overview", ["Item", "Description"], overview, [24, 110])

    # 2/3. Variable definitions for the columns actually present.
    node_cols = list(nodes[0].keys()) if nodes else []
    edge_cols = list(edges[0].keys()) if edges else []
    sheet("Node Attributes", ["Column", "Definition"],
          [(c, _col_def(c, _NODE_DEFS, _NODE_PREFIX_DEFS)) for c in node_cols], [28, 105])
    sheet("Edge Attributes", ["Column", "Definition"],
          [(c, _col_def(c, _EDGE_DEFS)) for c in edge_cols], [28, 105])

    # 4. Entity types with counts (+ domain subtypes where defined).
    subtype_map = {}
    if domain is not None:
        try:
            subtype_map = domain.entity_subtypes() or {}
        except Exception:  # noqa: BLE001
            subtype_map = {}
    type_rows = []
    for t, n in type_counts.most_common():
        subs = ", ".join(subtype_map.get(t, []))
        type_rows.append((t, n, f"Subtypes in tag_subtype: {subs}" if subs else ""))
    sheet("Entity Types", ["Type", "Nodes", "Notes"], type_rows, [16, 10, 100])

    # 5. Tie classes with counts.
    tie_rows = []
    for cls, definition in _TIE_CLASS_DEFS.items():
        role = ("social" if cls in tie_classes.SOCIAL else
                "structural" if cls in tie_classes.STRUCTURAL else "non-social")
        tie_rows.append((cls, class_counts.get(cls, 0), role, definition))
    sheet("Tie Classes", ["Class", "Edges", "Role", "Definition"], tie_rows, [16, 10, 12, 95])

    # 6. Relation inventory: every rel_type in this run with an example.
    rel_info: dict[str, dict] = {}
    for e in edges:
        rt = e.get("rel_type", "")
        info = rel_info.setdefault(rt, {"n": 0, "tie_class": e.get("tie_class", ""),
                                        "connection_type": e.get("connection_type", ""),
                                        "polarity": e.get("polarity", ""), "example": ""})
        info["n"] += 1
        if not info["example"] and e.get("evidence"):
            info["example"] = str(e["evidence"])[:200]
    rel_rows = [(rt, i["tie_class"], i["connection_type"], i["polarity"], i["n"], i["example"])
                for rt, i in sorted(rel_info.items(), key=lambda kv: -kv[1]["n"])]
    sheet("Relation Inventory",
          ["rel_type", "tie_class", "connection_type", "polarity", "Edges", "Example evidence"],
          rel_rows, [26, 14, 14, 10, 8, 80])

    # 7. Evidence tiers.
    source_counts = Counter(s for e in edges for s in str(e.get("edge_source", "")).split(";") if s)
    tier_rows = [(t, srcs, d) for t, srcs, d in TIER_DOCS]
    tier_rows.append(("", "", ""))
    tier_rows.append(("edge_source values in this run",
                      ", ".join(f"{s} ({n})" for s, n in source_counts.most_common()), ""))
    sheet("Evidence Tiers", ["Tier", "edge_source values", "Definition"], tier_rows, [22, 60, 60])

    wb.remove(wb["Sheet"])
    out = run_dir / "codebook.xlsx"
    wb.save(out)
    return out
